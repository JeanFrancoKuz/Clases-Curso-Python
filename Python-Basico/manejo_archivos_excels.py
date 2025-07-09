#¿Que es openpyxl?
#Es una libreria para trabajar con archivos excel

#Crear un archivo de Excel 

from openpyxl import load_workbook,Workbook

#Importar los estilos
from openpyxl.styles import Font

workbook = Workbook()
workbook.save("./Python-Basico/relatos.xlsx")

#seleccionarlo y agregarle valores

book = workbook.active

#Escribir datos

book["A1"] = "Producto"
book["B1"] = "Precio"
book["C1"] = "Cantidad"

#Valores

book["A2"] = "Camisa"
book["B2"] = "5000"
book["C2"] = "10"

#Para guardar cambios utilizamos el metodo save()

workbook.save("./Python-Basico/relatos.xlsx")

#para leer tendre la clase load_workbook

#Aqui solo cargo el archivo sin abrirlo

book = load_workbook("./Python-Basico/relatos.xlsx")

hoja = book.active

#Vamos a leer los datos

product = hoja["A2"].value
price = hoja["B2"].value
quantity = hoja["C2"].value

print(f"Producto: {product}, Precio: {price}, Cantidad: {quantity}") 

#Crear un libro que se llama ventas y cargarlo

book = Workbook()

#Crear una hoja
hoja = book.active

#Cambiar el titulo

hoja.title = "Ventas-Enero"

#encabezados

hoja["A1"] = "Dia"
hoja["B1"] = "Ventas (USD)"

book.save("./Python-Basico/ventas.xlsx")

hoja["A1"].font = Font(bold=True)
hoja["B1"].font = Font(bold=True)

#creemos valores

#Un array con varios arrays dentro se llama matriz
ventas = [
  ["Lunes", 100],
  ["Martes", 200],
  ["Miercoles", 300],
  ["Jueves", 400],
  ["Viernes", 500],
  ["Sabado", 600],
  ["Domingo", 700],
]

#Para insertar mutiples valores necesitaremos utilizar un bucle

for row in ventas:
  hoja.append(row) #agregara este valor automaticamente a la siguiente fila vacia
  
book.save("./Python-Basico/ventas.xlsx")

#leer el contenido

book = load_workbook("./Python-Basico/ventas.xlsx")

hoja = book.active

#mostrar mi contenido

for row in hoja.iter_rows(min_row=2):
  #Cada row se ve asi (celda1=dia, celda2=ventas)
  print(row[0].value, row[1].value)
  
#Ejercicio 

#Crea un archivo ventas_trimestrales y agrega los datos de la siguiente manera:
# [trimestre, ventas totales]

#luego ingresa los valores de todos los trimestres de un año

#Luego muestra las ventas totales de cada trimestre